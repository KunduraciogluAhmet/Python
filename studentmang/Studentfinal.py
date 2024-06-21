from tkinter import *
from datetime import date
from tkinter import filedialog, messagebox
from PIL import Image, ImageTk
import os
from tkinter.ttk import Combobox
import openpyxl
from openpyxl import Workbook
import pathlib

background = "#542b2a"
framebg = "#ededed"
framefg = "#06283d"

root = Tk()
root.title("Student Registration System")
root.geometry("1250x700+20+10")
root.resizable(0, 0)
root.config(bg=background)

file = pathlib.Path('Student_data.xlsx')
if not file.exists():
    workbook = Workbook()
    sheet = workbook.active
    sheet['A1'] = "Reg No"
    sheet['B1'] = "Name"
    sheet['C1'] = "Class"
    sheet['D1'] = "Gender"
    sheet['E1'] = "DOB"
    sheet['F1'] = "Date of Reg"
    sheet['G1'] = "Religion"
    sheet['H1'] = "Skill"
    sheet['I1'] = "Father's Name"
    sheet['J1'] = "Mother's Name"
    sheet['K1'] = "Father's Occupation"
    sheet['L1'] = "Mother's Occupation"
    sheet['M1'] = "Status"
    workbook.save("Student_data.xlsx")

def Exit():
    root.destroy()

def showImage():
    filename = filedialog.askopenfilename(initialdir=os.getcwd(), title="Select image file", filetypes=(("jpg File", "*.jpg"), ("jpeg File", "*.jpeg"), ("PNG File", "*.png"), ("All Files", "*.*")))
    if filename:
        img = Image.open(filename)
        resized_image = img.resize((190, 190))
        photo2 = ImageTk.PhotoImage(resized_image)
        lbl.config(image=photo2)
        lbl.image = photo2
        global img_path
        img_path = filename

def registration_no():
    file = openpyxl.load_workbook('Student_data.xlsx')
    sheet = file.active
    row = sheet.max_row
    max_row_value = sheet.cell(row=row, column=1).value
    try:
        Registration.set(max_row_value + 1)
    except:
        Registration.set("1")

def Clear():
    global img
    Name.set('')
    DOB.set('')
    Religion.set('')
    Skill.set('')        
    F_Name.set('')
    M_Name.set('')
    F_Occup.set('')
    M_Occup.set('')  
    Class.set("Select Class")
    registration_no()
    saveButton.config(state="normal")
    img1 = PhotoImage(file="Images/uploadphoto.png")
    lbl.config(image=img1)
    lbl.image = img1
    img = ""

def Save():
    R1 = Registration.get()
    N1 = Name.get()
    C1 = Class.get()
    try:
        G1 = gender
    except:
        messagebox.showerror("Error", "Select Gender!")
    
    D2 = DOB.get()
    D1 = Date.get()
    Re1 = Religion.get()
    S1 = Skill.get()
    fathername = F_Name.get()
    mothername = M_Name.get()
    F1 = F_Occup.get()
    M1 = M_Occup.get()

    if N1 == "" or C1 == "Select Class" or D2 == "" or Re1 == "" or S1 == "" or  fathername == "" or mothername == "" or F1 == "" or M1 == "":
        messagebox.showerror("ERROR", "Missing Data")
    else:
        file = openpyxl.load_workbook("Student_data.xlsx")
        sheet = file.active
        sheet.append([R1, N1, C1, G1, D2, D1, Re1, S1, fathername, mothername, F1, M1])

        file.save(r'Student_data.xlsx')

        try:
            img.save("StudentImages/" + str(R1) + ".jpg")
        except:
            messagebox.showinfo("INFO", "Profile Picture is Not Available!")
        
        messagebox.showinfo("INFO", "Successfully Recorded")
        Clear()
        registration_no()

def selection():
    global gender
    value = radio.get()
    if value == 1:
        gender = "Male"
    else:
        gender = "Female"

def Search():
    text = SearchVar.get()
    Clear()
    saveButton.config(state='disable')

    file = openpyxl.load_workbook("Student_data.xlsx")
    sheet = file.active

    found = False

    for row in sheet.iter_rows():
        if str(row[0].value) == text:
            found = True
            Registration.set(row[0].value)
            Name.set(row[1].value)
            Class.set(row[2].value)
            gender = row[3].value
            DOB.set(row[4].value)
            Date.set(row[5].value)
            Religion.set(row[6].value)
            Skill.set(row[7].value)
            F_Name.set(row[8].value)
            M_Name.set(row[9].value)
            F_Occup.set(row[10].value)
            M_Occup.set(row[11].value)

            if gender == "Male":
                radio.set(1)
            elif gender == "Female":
                radio.set(2)

            img_path = f"StudentImages/{row[0].value}.jpg"
            if os.path.exists(img_path):
                img = Image.open(img_path)
                resized_image = img.resize((190, 190))
                photo2 = ImageTk.PhotoImage(resized_image)
                lbl.config(image=photo2)
                lbl.image = photo2
            else:
                messagebox.showinfo("INFO", "Profile picture not found!")

            break

    if not found:
        messagebox.showerror("Invalid Query", "Record not found")

Label(root, text="E-Mail: ahmettkunduracioglu@gmail.com", width=10, height=2, bg="#e1798d", anchor="e").pack(side=BOTTOM, fill=X)
Label(root, text="STUDENT REGISTRATION", width=10, height=2, bg="#c36464", fg="#fff", font="arial 20 bold").pack(side=TOP, fill=X)

SearchVar = StringVar()
Entry(root, textvariable=SearchVar, width=15, bd=2, font="arial 20").place(x=820, y=16)
imageicon3 = PhotoImage(file="Images/search.png")
Srch = Button(root, text="Search", compound=LEFT, image=imageicon3, width=123, bg='#68ddfa', font="arial 13 bold")
Srch.place(x=1060, y=15)

Srch.config(command=Search)

imageicon4 = PhotoImage(file="Images/layer4.png")
Update_button = Button(root, image=imageicon4, bg="#c36464")
Update_button.place(x=110, y=15)

Label(root, text="Reg. No:", font="arial 13", fg=framefg, bg=background).place(x=30, y=80)
Label(root, text="Date:", font="arial 13", fg=framefg, bg=background).place(x=500, y=80)

Registration = IntVar()
Date = StringVar()

reg_entry = Entry(root, textvariable=Registration, width=15, font="arial 10")
reg_entry.place(x=160, y=80)

registration_no()

today = date.today()
d1 = today.strftime("%d/%m/%Y")
date_entry = Entry(root, textvariable=Date, width=15, font="arial 10")
date_entry.place(x=550, y=80)

Date.set(d1)

# Student Details
obj = LabelFrame(root, text="Student's Details", font=20, bd=2, width=900, bg=framebg, fg=framefg, height=250, relief=GROOVE)
obj.place(x=30, y=125)

Label(obj, text="Full Name:", font="arial 13", bg=framebg, fg=framefg).place(x=20, y=15)
Label(obj, text="Date of Birth:", font="arial 13", bg=framebg, fg=framefg).place(x=20, y=65)
Label(obj, text="Gender:", font="arial 13", bg=framebg, fg=framefg).place(x=20, y=115)
Label(obj, text="Class:", font="arial 13", bg=framebg, fg=framefg).place(x=500, y=15)
Label(obj, text="Religion:", font="arial 13", bg=framebg, fg=framefg).place(x=500, y=65)
Label(obj, text="Skills:", font="arial 13", bg=framebg, fg=framefg).place(x=500, y=115)

Name = StringVar()
DOB = StringVar()
Religion = StringVar()
Skill = StringVar()
F_Name = StringVar()
M_Name = StringVar()
F_Occup = StringVar()
M_Occup = StringVar()
Class = StringVar()

name_entry = Entry(obj, textvariable=Name, width=20, font="arial 10")
name_entry.place(x=160, y=15)

dob_entry = Entry(obj, textvariable=DOB, width=20, font="arial 10")
dob_entry.place(x=160, y=65)

radio = IntVar()
R1 = Radiobutton(obj, text="Male", variable=radio, value=1, bg=framebg, fg=framefg, command=selection)
R1.place(x=150, y=115)

R2 = Radiobutton(obj, text="Female", variable=radio, value=2, bg=framebg, fg=framefg, command=selection)
R2.place(x=200, y=115)

Class.set("Select Class")
class_entry = OptionMenu(obj, Class, "1", "2", "3", "4", "5", "6", "7", "8", "9", "10", "11", "12")
class_entry.place(x=630, y=11)
class_entry.config(width=15)

religion_entry = Entry(obj, textvariable=Religion, width=20, font="arial 10")
religion_entry.place(x=630, y=65)

skill_entry = Entry(obj, textvariable=Skill, width=20, font="arial 10")
skill_entry.place(x=630, y=115)

# Parents Details
obj2 = LabelFrame(root, text="Parent's Details", font=20, bd=2, width=900, bg=framebg, fg=framefg, height=220, relief=GROOVE)
obj2.place(x=30, y=400)

Label(obj2, text="Father's Name:", font="arial 13", bg=framebg, fg=framefg).place(x=20, y=15)
Label(obj2, text="Occupation:", font="arial 13", bg=framebg, fg=framefg).place(x=20, y=65)

f_entry = Entry(obj2, textvariable=F_Name, width=20, font="arial 10")
f_entry.place(x=160, y=15)

fo_entry = Entry(obj2, textvariable=F_Occup, width=20, font="arial 10")
fo_entry.place(x=160, y=65)

Label(obj2, text="Mother's Name:", font="arial 13", bg=framebg, fg=framefg).place(x=500, y=15)
Label(obj2, text="Occupation:", font="arial 13", bg=framebg, fg=framefg).place(x=500, y=65)

m_entry = Entry(obj2, textvariable=M_Name, width=20, font="arial 10")
m_entry.place(x=630, y=15)

mo_entry = Entry(obj2, textvariable=M_Occup, width=20, font="arial 10")
mo_entry.place(x=630, y=65)

# Image
f = Frame(root, bd=3, bg="black", width=200, height=200, relief=GROOVE)
f.place(x=1000, y=150)

img = PhotoImage(file="Images/uploadphoto.png")
lbl = Label(f, bg="black", image=img)
lbl.place(x=0, y=0)

# Button
Button(root, text="Upload", width=19, height=2, font="arial 12 bold", bg="lightblue", command=showImage).place(x=1000, y=370)
saveButton = Button(root, text="Save", width=19, height=2, font="arial 12 bold", bg="lightgreen", command=Save)
saveButton.place(x=1000, y=450)
Button(root, text="Reset", width=19, height=2, font="arial 12 bold", bg="lightpink", command=Clear).place(x=1000, y=530)
Button(root, text="Exit", width=19, height=2, font="arial 12 bold", bg="grey", command=Exit).place(x=1000, y=610)

root.mainloop()
